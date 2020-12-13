import openpyxl

def getRowCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return (sheet.max_row)

def getColCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rowNo,colNo):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.cell(row=rowNo,column=colNo).value

def writeData(file,sheetName,rowNo,colNo,data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.cell(row=rowNo,column=colNo).value=data
    wb.save(file)

